#!/usr/bin/env python3
"""
export_prompts.py — Export all pipeline prompts to a structured Excel file.

Tabs:
  1. global_opinion_qa  — original question | final prompt sent
  2. persona_prompts    — one row per query; baseline | high_ses | low_ses columns

Usage:
  python export_prompts.py
  python export_prompts.py --out results/prompts_overview.xlsx
"""

import argparse
import json
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── Helpers ───────────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(bold=True, color="FFFFFF")
ALT_FILL      = PatternFill("solid", fgColor="D6E4F0")
WRAP_ALIGN    = Alignment(wrap_text=True, vertical="top")


def _style_sheet(ws, col_widths: list[int]) -> None:
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for cell in ws[1]:
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for cell in row:
            cell.alignment = WRAP_ALIGN
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


# ── Tab 1: Global Opinion QA ──────────────────────────────────────────────────

def build_goqa(path: Path) -> pd.DataFrame:
    items = json.loads(path.read_text())
    rows = []
    for it in items:
        rows.append({
            "item_id":           it["item_id"],
            "source":            it.get("source", "WVS"),
            "entropy":           round(it.get("entropy", 0.0), 4),
            "is_binary":         it.get("is_binary", False),
            "original_question": it["question"],
            "prompt_sent":       it["prompt"],
        })
    return pd.DataFrame(rows)


# ── Tab 2: Persona Prompts ────────────────────────────────────────────────────

def build_persona(path: Path) -> pd.DataFrame:
    items = json.loads(path.read_text())

    # Index by (query_source, query_id) → {condition: item}
    lookup: dict[tuple, dict] = {}
    for it in items:
        key = (it["query_source"], it["query_id"])
        lookup.setdefault(key, {})[it["condition"]] = it

    rows = []
    for (source, qid), conds in sorted(lookup.items(), key=lambda x: (x[0][0], x[0][1])):
        baseline = conds.get("baseline", {})
        high     = conds.get("high_ses", {})
        low      = conds.get("low_ses",  {})

        rows.append({
            "query_source":      source,
            "query_id":          qid,
            "baseline_prompt":   baseline.get("prompt", ""),

            # High SES
            "high_ses_persona":  high.get("persona_text", ""),
            "high_ses_format":   high.get("persona_role", ""),   # "customer" or "user"
            "high_ses_system":   high.get("system") or "",
            "high_ses_user_msg": high.get("prompt", ""),

            # Low SES
            "low_ses_persona":   low.get("persona_text", ""),
            "low_ses_format":    low.get("persona_role", ""),
            "low_ses_system":    low.get("system") or "",
            "low_ses_user_msg":  low.get("prompt", ""),
        })

    return pd.DataFrame(rows)


# ── Write Excel ───────────────────────────────────────────────────────────────

def write_excel(goqa_df: pd.DataFrame, persona_df: pd.DataFrame, out: Path) -> None:
    out.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        # ── Tab 1 ──────────────────────────────────────────────────────────────
        goqa_df.to_excel(writer, sheet_name="global_opinion_qa", index=False)
        ws1 = writer.sheets["global_opinion_qa"]
        _style_sheet(ws1, col_widths=[22, 10, 10, 10, 60, 80])

        # ── Tab 2 ──────────────────────────────────────────────────────────────
        persona_df.to_excel(writer, sheet_name="persona_prompts", index=False)
        ws2 = writer.sheets["persona_prompts"]
        _style_sheet(ws2, col_widths=[18, 10, 45, 45, 12, 45, 45, 45, 12, 45, 45])

        HIGH_FILL = PatternFill("solid", fgColor="E8F4E8")
        LOW_FILL  = PatternFill("solid", fgColor="FFF0E8")
        high_cols = [4, 5, 6, 7]
        low_cols  = [8, 9, 10, 11]
        for row in ws2.iter_rows(min_row=2):
            for cell in row:
                if cell.column in high_cols:
                    cell.fill = HIGH_FILL
                elif cell.column in low_cols:
                    cell.fill = LOW_FILL

    print(f"Saved → {out}")
    print(f"  global_opinion_qa : {len(goqa_df)} rows")
    print(f"  persona_prompts   : {len(persona_df)} rows  ({len(persona_df) * 3} total prompts incl. 3 conditions)")


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--goqa",    default="globalopinionqa_wvs.json")
    p.add_argument("--persona", default="persona_prompts.json")
    p.add_argument("--out",     default="results/prompts_overview.xlsx")
    args = p.parse_args()

    goqa_df    = build_goqa(Path(args.goqa))
    persona_df = build_persona(Path(args.persona))

    write_excel(goqa_df, persona_df, Path(args.out))


if __name__ == "__main__":
    main()
